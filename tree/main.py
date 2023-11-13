import openpyxl


class Tree:
    def __init__(self, excel_name, *classifiers):
        self.depth = len(classifiers)  # Глубина дерева, она же количество классификаторов
        self.classifiers = {}  # Словарь имя_классификатора - индекс_классификатора
        self.objects = {}  # Словарь значения_классификаторов - объекты
        # Список словарей значение_классификатора - индекс_значения
        self.class_values = [{} for _ in range(self.depth)]
        for i, classifier in enumerate(classifiers):
            self.classifiers[classifier] = i

        indexes = [0] * self.depth
        sheet = openpyxl.load_workbook(excel_name).active
        for i in range(1, sheet.max_column):
            classifier_name = sheet[1][i].value
            if classifier_name in self.classifiers:
                indexes[self.classifiers[classifier_name]] = i
        for row in range(2, sheet.max_row + 1):
            obj_name = sheet[row][0].value
            key = []
            for i, idx in enumerate(indexes):
                classifier_val = sheet[row][idx].value
                if classifier_val not in self.class_values[i]:
                    self.class_values[i][classifier_val] = len(self.class_values[i])
                key.append(self.class_values[i][classifier_val])
            key = tuple(key)
            if key not in self.objects:
                self.objects[key] = set()
            self.objects[key].add(obj_name)
        # Многомерный список, хранящий объекты
        self.root = self.new_vertex([])
        # Словарь ограничений для get_children; индекс_классификатора - индекс_значения
        self.encoded_restrictions = {}

    def new_vertex(self, stack):
        # Рекурсивный метод формирования дерева
        """
        :param stack: стек индексов значений классификаторов
        :return: список детей (объекты, если это последний классификатор)
        """
        children = []  # Список вершин-детей, индекс - индекс значения соотв. классификатора
        depth = len(stack)
        if depth != self.depth - 1:
            # Стек дополняется индексами возможных значений следующего классификатора
            for val in self.class_values[depth].values():
                children.append(self.new_vertex(stack + [val]))
            return children
        # Если вершина листовая - в качестве детей идут объекты
        for val in self.class_values[depth].values():
            key = tuple(stack + [val])
            if key in self.objects:
                children.append(self.objects[key])
            else:
                children.append(set())
        return children

    def get_children(self, *restrictions):
        # Кодировка ограничений (сначала ограничения обнуляются)
        self.encoded_restrictions = {}
        for restriction in restrictions:
            classifier_idx = self.classifiers[restriction[0]]
            value_idx = self.class_values[classifier_idx][restriction[1]]
            self.encoded_restrictions[classifier_idx] = value_idx
        res = self.children(0, self.root)
        return res

    def children(self, depth, next_level):
        """
        :param depth: текущая глубина (индекс классификатора)
        :param next_level: список детей
        :return: объекты, удовлетворяющие ограничениям в данном поддереве
        """
        # Рекурсивный метод спуска по дереву
        objects = set()
        if depth != self.depth - 1:
            if depth in self.encoded_restrictions:
                objects |= self.children(depth + 1, next_level[self.encoded_restrictions[depth]])
            else:
                for i in range(len(next_level)):
                    objects |= self.children(depth + 1, next_level[i])
            return objects

        if depth in self.encoded_restrictions:
            return next_level[self.encoded_restrictions[depth]]
        for i in next_level:
            objects |= i
        return objects


def test():
    test_tree = Tree("test.xlsx", "classifier1", "classifier2", "classifier3")
    assert set() == test_tree.get_children(("classifier3", 31), ("classifier1", 12), ("classifier2", 21))
    assert {"name2"} == test_tree.get_children(("classifier3", 31), ("classifier1", 12), ("classifier2", 22))
    assert {"name2", "name6"} == test_tree.get_children(("classifier3", 31), ("classifier1", 12))
    assert {"name1", "name4", "name7"} == test_tree.get_children(("classifier2", 21))
    assert {"name1", "name2", "name3", "name4", "name5", "name6", "name7"} == test_tree.get_children()


test()
